#!/usr/bin/env python3
"""Summarize MCXN947 FlexIO pin candidates against the FRDM workbook."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SIGNAL_JSON = (
    REPO_ROOT / "docs" / "research" / "MCXN947" / "json" / "ksdk2_0" / "MCXN947VDF" / "signal_configuration.json"
)
DEFAULT_WORKBOOK = (
    REPO_ROOT
    / "docs"
    / "FRDM-MCXN947"
    / "FRDM-MCXN947"
    / "FRDM-MCXN947 board pin assignment.xlsx"
)
DEFAULT_OUTPUT = REPO_ROOT / "docs" / "research" / "MCXN947" / "flexio_pin_candidates"
ALLOWED_CLEAN_ROOT = REPO_ROOT / "docs" / "research" / "MCXN947"


@dataclass(frozen=True)
class FlexioPin:
    pio: str
    port: int
    pin: int
    bga_coord: str
    flexio_data: int
    mux_alt: str
    route_signals: str
    pin_name: str
    board_primary_assignment: str
    board_on_board_pin: str
    board_description: str
    board_summary_1: str
    board_summary_2: str
    board_summary_3: str
    board_summary_4: str
    board_flexio_lcd: str
    board_ezh_camera: str
    board_arduino: str
    board_mikroe: str
    board_pmod: str


@dataclass(frozen=True)
class HeaderSignal:
    source: str
    header_pin: str
    logical_signal: str
    pio: str
    pin_description: str
    workbook_flexio_data: str
    nxp_flexio_data: str
    status: str
    note: str


def repo_relative(path: Path) -> str:
    try:
        return path.resolve().relative_to(REPO_ROOT).as_posix()
    except ValueError:
        return str(path.resolve())


def as_list(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def extract_pio(value: Any) -> str:
    match = re.search(r"\b(?:PIO|P)(\d+)_(\d+)\b", text(value))
    if not match:
        return ""
    return f"P{match.group(1)}_{match.group(2)}"


def parse_pio(pio: str) -> tuple[int, int]:
    match = re.fullmatch(r"P(\d+)_(\d+)", pio)
    if not match:
        raise ValueError(f"invalid PIO name: {pio}")
    return int(match.group(1)), int(match.group(2))


def extract_workbook_flexio_data(value: Any) -> str:
    match = re.search(r"\b(?:FLEXIO0_|FXIO_?)D(\d+)(?!\d)", text(value), flags=re.IGNORECASE)
    if not match:
        return ""
    return f"D{int(match.group(1))}"


def walk_dicts(value: Any) -> list[dict[str, Any]]:
    found: list[dict[str, Any]] = []
    if isinstance(value, dict):
        found.append(value)
        for child in value.values():
            found.extend(walk_dicts(child))
    elif isinstance(value, list):
        for child in value:
            found.extend(walk_dicts(child))
    return found


def find_mux_alt(configuration: Any, pio: str) -> str:
    port, pin = parse_pio(pio)
    register = f"PORT{port}_PCR{pin}"
    for item in walk_dicts(configuration):
        if item.get("@register") != register or item.get("@bit_field") != "MUX":
            continue
        raw_value = text(item.get("@bit_field_value"))
        if not raw_value:
            continue
        try:
            return f"ALT{int(raw_value, 0)}"
        except ValueError:
            return raw_value
    return ""


def read_board_annotations(workbook_path: Path) -> dict[str, dict[str, str]]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["MCXN947_Pinmux"]
    rows = ws.iter_rows(values_only=True)
    headers = [text(value).replace("\n", " ") for value in next(rows)]
    annotations: dict[str, dict[str, str]] = {}
    for row in rows:
        values = {headers[index]: text(value) for index, value in enumerate(row) if index < len(headers)}
        pio = values.get("184BGA ALL Pin Name", "")
        if pio:
            annotations[pio] = values
    return annotations


def read_flexio_pins(signal_json: Path, board_annotations: dict[str, dict[str, str]]) -> list[FlexioPin]:
    data = json.loads(signal_json.read_text(encoding="utf-8"))
    pins = data["signal_configuration"]["pins"]["pin"]
    flexio_pins: list[FlexioPin] = []

    for pin in pins:
        pin_name = text(pin.get("@name"))
        pio = extract_pio(pin_name)
        if not pio:
            continue
        flexio_matches = re.findall(r"\bFLEXIO0_D(\d+)\b", pin_name)
        if not flexio_matches:
            continue

        port, pin_number = parse_pio(pio)
        route_descriptions: list[str] = []
        mux_alt = ""

        for outer in as_list(pin.get("connections")):
            connection = outer.get("connection") if isinstance(outer, dict) else None
            if not isinstance(connection, dict):
                continue
            signal_ref = connection.get("peripheral_signal_ref")
            if not isinstance(signal_ref, dict) or signal_ref.get("@peripheral") != "FLEXIO0":
                continue
            signal = text(signal_ref.get("@signal"))
            channel = text(signal_ref.get("@channel"))
            route_descriptions.append(f"{signal}{channel}".strip())
            mux_alt = mux_alt or find_mux_alt(connection.get("configuration"), pio)

        annotations = board_annotations.get(pio, {})
        for flexio_match in flexio_matches:
            flexio_pins.append(
                FlexioPin(
                    pio=pio,
                    port=port,
                    pin=pin_number,
                    bga_coord=text(pin.get("@coords")),
                    flexio_data=int(flexio_match),
                    mux_alt=mux_alt,
                    route_signals=", ".join(sorted(set(route_descriptions))),
                    pin_name=pin_name,
                    board_primary_assignment=annotations.get("Primary Assignment", ""),
                    board_on_board_pin=annotations.get("On-board pin", ""),
                    board_description=annotations.get("Description", ""),
                    board_summary_1=annotations.get("Summary(1st option)", ""),
                    board_summary_2=annotations.get("(2nd option)", ""),
                    board_summary_3=annotations.get("(3rd option)", ""),
                    board_summary_4=annotations.get("(4th option)", ""),
                    board_flexio_lcd=annotations.get("FlexIO/LCD", ""),
                    board_ezh_camera=annotations.get("EZH/Camera", ""),
                    board_arduino=annotations.get("Arduino", ""),
                    board_mikroe=annotations.get("MikroE", ""),
                    board_pmod=annotations.get("PMOD", ""),
                )
            )

    flexio_pins.sort(key=lambda item: (item.port, item.pin, item.flexio_data))
    return flexio_pins


def flexio_by_pio(flexio_pins: list[FlexioPin]) -> dict[str, list[FlexioPin]]:
    result: dict[str, list[FlexioPin]] = {}
    for pin in flexio_pins:
        result.setdefault(pin.pio, []).append(pin)
    return result


def annotate_header_signal(
    source: str,
    logical_signal: Any,
    pin_description: Any,
    header_pin: Any,
    pio_to_flexio: dict[str, list[FlexioPin]],
) -> HeaderSignal | None:
    logical = text(logical_signal)
    pin_desc = text(pin_description)
    header = text(header_pin)
    if not logical and not pin_desc:
        return None
    if not pin_desc and logical.upper() in {"CAMERA", "LCD", "EZH", "FLEXIO HEADER", "FLEXIO CAMERA HEADER"}:
        return None

    pio = extract_pio(pin_desc)
    workbook_flexio_data = extract_workbook_flexio_data(pin_desc)
    nxp_data = ",".join(f"D{pin.flexio_data}" for pin in pio_to_flexio.get(pio, []))

    if not pio:
        status = "not_mcu_pin"
        note = "Power, ground, or non-PIO entry."
    elif workbook_flexio_data and not nxp_data:
        status = "workbook_nxp_mismatch"
        note = "Workbook labels this as FlexIO, but MCXN947VDF signal data does not."
    elif workbook_flexio_data and workbook_flexio_data not in nxp_data.split(","):
        status = "workbook_nxp_mismatch"
        note = f"Workbook says {workbook_flexio_data}; NXP signal data says {nxp_data or 'no FlexIO'}."
    elif workbook_flexio_data:
        status = "matches_nxp"
        note = "Workbook FlexIO label matches MCXN947VDF signal data."
    elif nxp_data:
        status = "nxp_flexio_available"
        note = "NXP signal data exposes FlexIO on this pin; workbook entry is not labeled as FlexIO."
    else:
        status = "not_flexio"
        note = "No FlexIO function expected for this signal."

    return HeaderSignal(
        source=source,
        header_pin=header,
        logical_signal=logical,
        pio=pio,
        pin_description=pin_desc,
        workbook_flexio_data=workbook_flexio_data,
        nxp_flexio_data=nxp_data,
        status=status,
        note=note,
    )


def read_ezh_camera_header(workbook_path: Path, pio_to_flexio: dict[str, list[FlexioPin]]) -> list[HeaderSignal]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["EZH_CAMERA"]
    records: list[HeaderSignal] = []

    for column in range(2, ws.max_column + 1):
        top = annotate_header_signal(
            "EZH_CAMERA top",
            ws.cell(6, column).value,
            ws.cell(8, column).value,
            ws.cell(9, column).value,
            pio_to_flexio,
        )
        if top:
            records.append(top)

        bottom = annotate_header_signal(
            "EZH_CAMERA bottom",
            ws.cell(13, column).value,
            ws.cell(11, column).value,
            ws.cell(10, column).value,
            pio_to_flexio,
        )
        if bottom:
            records.append(bottom)

    return records


def read_flexio_camera_a18_header(workbook_path: Path, pio_to_flexio: dict[str, list[FlexioPin]]) -> list[HeaderSignal]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["FLEXIO_LCD A18"]
    records: list[HeaderSignal] = []

    for row in range(9, ws.max_row + 1):
        left = annotate_header_signal(
            "FLEXIO_LCD A18 camera left",
            ws.cell(row, 7).value,
            ws.cell(row, 8).value,
            ws.cell(row, 9).value,
            pio_to_flexio,
        )
        if left:
            records.append(left)

        right = annotate_header_signal(
            "FLEXIO_LCD A18 camera right",
            ws.cell(row, 12).value,
            ws.cell(row, 11).value,
            ws.cell(row, 10).value,
            pio_to_flexio,
        )
        if right:
            records.append(right)

    return records


def read_flexio_lcd_header(workbook_path: Path, pio_to_flexio: dict[str, list[FlexioPin]]) -> list[HeaderSignal]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["FLEXIO_LCD"]
    records: list[HeaderSignal] = []

    for row in range(9, ws.max_row + 1):
        left = annotate_header_signal(
            "FLEXIO_LCD left",
            "",
            ws.cell(row, 2).value,
            ws.cell(row, 3).value,
            pio_to_flexio,
        )
        if left:
            records.append(left)

        right = annotate_header_signal(
            "FLEXIO_LCD right",
            "",
            ws.cell(row, 5).value,
            ws.cell(row, 4).value,
            pio_to_flexio,
        )
        if right:
            records.append(right)

    return records


def write_csv(path: Path, rows: list[Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    dict_rows = [asdict(row) for row in rows]
    if not dict_rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=list(dict_rows[0].keys()))
        writer.writeheader()
        writer.writerows(dict_rows)


def write_json(path: Path, rows: list[Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps([asdict(row) for row in rows], indent=2, sort_keys=True) + "\n", encoding="utf-8")


def markdown_table(headers: list[str], rows: list[list[str]]) -> list[str]:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(cell.replace("|", "\\|") for cell in row) + " |")
    return lines


def write_markdown(
    path: Path,
    signal_json: Path,
    workbook_path: Path,
    flexio_pins: list[FlexioPin],
    ezh_camera: list[HeaderSignal],
    flexio_camera_a18: list[HeaderSignal],
    flexio_lcd_header: list[HeaderSignal],
) -> None:
    port4_flexio = [pin for pin in flexio_pins if pin.port == 4]
    valid_flexio_header = [
        record
        for record in flexio_lcd_header
        if record.status in {"matches_nxp", "nxp_flexio_available"} and record.pio.startswith("P4_")
    ]
    a18_mismatches = [record for record in flexio_camera_a18 if record.status == "workbook_nxp_mismatch"]
    current_camera = [
        record
        for record in ezh_camera
        if record.logical_signal
        and record.logical_signal
        not in {"3V3", "GND"}
    ]

    candidate_rows = [
        ["D0-D7", "P4_12-P4_19", "FLEXIO0_D20-D27", "Use contiguous Port 4 FlexIO pins; requires camera data fly-wires."],
        ["PCLK", "P4_20", "FLEXIO0_D28", "Route returned camera pixel clock here for first FlexIO timer tests."],
        ["HSYNC/HREF", "P4_21", "FLEXIO0_D29", "Use as FlexIO input if hardware gating is needed; otherwise GPIO/PINT can count it."],
        ["VSYNC", "P4_22 or GPIO/PINT", "FLEXIO0_D30 optional", "Start with GPIO/PINT instrumentation unless FlexIO hardware gating needs it."],
        ["Debug spare", "P4_23", "FLEXIO0_D31", "Keep as spare scope/debug input or output."],
        ["SCCB/I2C", "P3_2/P3_3", "not FlexIO path", "Keep current LP_FLEXCOMM7 wiring for first tests."],
        ["XCLK", "P2_2", "CLKOUT, also FLEXIO0_D10", "Keep current SCG CLKOUT camera clock output."],
        ["RESET/PWDN", "P1_19/P1_18", "GPIO", "Keep current camera control pins."],
    ]

    lines = [
        "# FlexIO Camera Pin Candidates",
        "",
        "Generated by `scripts/tools/summarize_flexio_pins.py`.",
        "",
        "## Sources",
        "",
        f"- NXP signal dump: `{repo_relative(signal_json)}`",
        f"- FRDM workbook: `{repo_relative(workbook_path)}`",
        "- Current firmware pin mux: `src/avc/avc_core0/board/pin_mux.c`",
        "- Current camera setup: `src/avc/avc_core0/source/avc_io/bv_camera__interface.c`",
        "",
        "## Initial Conclusion",
        "",
        "Use the NXP signal dump as the authority for MCXN947VDF FlexIO routing. The FRDM workbook is still useful for",
        "connector intent, but the `FLEXIO_LCD A18` camera section has FlexIO labels that do not match the current",
        "MCXN947VDF signal data for `P4_2..P4_7`.",
        "",
        "The best first FlexIO camera candidate is the valid Port 4 FlexIO group on `P4_12..P4_23`, which provides",
        "`FLEXIO0_D20..D31`. Keep the existing SCCB/I2C pins, XCLK output, reset, and power-down pins for initial tests.",
        "",
        "## Candidate Wiring",
        "",
        *markdown_table(["Camera signal", "MCU pins", "FlexIO role", "Note"], candidate_rows),
        "",
        "## Current EZH Camera Header Map",
        "",
        *markdown_table(
            ["Signal", "Header pin", "MCU pin", "Workbook/function"],
            [
                [
                    record.logical_signal,
                    record.header_pin,
                    record.pio,
                    record.pin_description,
                ]
                for record in current_camera
            ],
        ),
        "",
        "## NXP-Confirmed Port 4 FlexIO Pins",
        "",
        *markdown_table(
            ["MCU pin", "FlexIO data", "BGA", "Mux", "Workbook board note"],
            [
                [
                    pin.pio,
                    f"D{pin.flexio_data}",
                    pin.bga_coord,
                    pin.mux_alt,
                    pin.board_description or pin.board_flexio_lcd or pin.board_primary_assignment,
                ]
                for pin in port4_flexio
            ],
        ),
        "",
        "## FRDM FlexIO Header Pins",
        "",
        *markdown_table(
            ["Header pin", "MCU pin", "Workbook text", "NXP FlexIO", "Status"],
            [
                [
                    record.header_pin,
                    record.pio,
                    record.pin_description,
                    record.nxp_flexio_data,
                    record.status,
                ]
                for record in valid_flexio_header
            ],
        ),
        "",
        "## Workbook A18 Camera Map Mismatches",
        "",
    ]

    if a18_mismatches:
        lines.extend(
            markdown_table(
                ["Signal", "Header pin", "MCU pin", "Workbook FlexIO", "NXP FlexIO", "Note"],
                [
                    [
                        record.logical_signal,
                        record.header_pin,
                        record.pio,
                        record.workbook_flexio_data,
                        record.nxp_flexio_data or "none",
                        record.note,
                    ]
                    for record in a18_mismatches
                ],
            )
        )
    else:
        lines.append("No mismatches found.")

    lines.extend(
        [
            "",
            "## Open Checks",
            "",
            "- Verify the physical header and solder-jumper state before wiring `P4_12..P4_23`.",
            "- Confirm whether the first FlexIO firmware path can use `dataPinStartIdx = 20` for D0-D7 on FlexIO D20-D27.",
            "- Start VSYNC/HSYNC as GPIO/PINT counters, then move either signal into FlexIO only if timer gating requires it.",
            "- Scope PCLK at the camera module and at the selected MCU pin before enabling DMA.",
            "",
            "## Generated Files",
            "",
            "- `flexio0_pins.csv` and `flexio0_pins.json`: all NXP FlexIO0-capable pins with workbook annotations.",
            "- `ezh_camera_header.csv`: current EZH camera workbook map.",
            "- `flexio_camera_a18_header.csv`: workbook A18 camera header map with NXP mismatch status.",
            "- `flexio_lcd_header.csv`: FRDM FlexIO header workbook map with NXP status.",
            "",
        ]
    )

    path.write_text("\n".join(lines), encoding="utf-8")


def clean_output(output_dir: Path) -> None:
    resolved_output = output_dir.resolve()
    allowed_root = ALLOWED_CLEAN_ROOT.resolve()
    if resolved_output == allowed_root or allowed_root not in resolved_output.parents:
        raise SystemExit(f"Refusing to clean output outside {allowed_root}: {resolved_output}")
    if resolved_output.exists():
        shutil.rmtree(resolved_output)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--signal-json", type=Path, default=DEFAULT_SIGNAL_JSON, help="NXP signal_configuration.json")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK, help="FRDM board pin assignment workbook")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output directory")
    parser.add_argument("--keep-existing", action="store_true", help="Do not clean the output directory before writing")
    args = parser.parse_args()

    signal_json = args.signal_json.resolve()
    workbook_path = args.workbook.resolve()
    output_dir = args.output.resolve()
    if not signal_json.exists():
        raise SystemExit(f"signal JSON not found: {signal_json}")
    if not workbook_path.exists():
        raise SystemExit(f"workbook not found: {workbook_path}")

    if not args.keep_existing:
        clean_output(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    board_annotations = read_board_annotations(workbook_path)
    flexio_pins = read_flexio_pins(signal_json, board_annotations)
    pio_to_flexio = flexio_by_pio(flexio_pins)
    ezh_camera = read_ezh_camera_header(workbook_path, pio_to_flexio)
    flexio_camera_a18 = read_flexio_camera_a18_header(workbook_path, pio_to_flexio)
    flexio_lcd_header = read_flexio_lcd_header(workbook_path, pio_to_flexio)

    write_csv(output_dir / "flexio0_pins.csv", flexio_pins)
    write_json(output_dir / "flexio0_pins.json", flexio_pins)
    write_csv(output_dir / "ezh_camera_header.csv", ezh_camera)
    write_json(output_dir / "ezh_camera_header.json", ezh_camera)
    write_csv(output_dir / "flexio_camera_a18_header.csv", flexio_camera_a18)
    write_json(output_dir / "flexio_camera_a18_header.json", flexio_camera_a18)
    write_csv(output_dir / "flexio_lcd_header.csv", flexio_lcd_header)
    write_json(output_dir / "flexio_lcd_header.json", flexio_lcd_header)
    write_markdown(
        output_dir / "README.md",
        signal_json,
        workbook_path,
        flexio_pins,
        ezh_camera,
        flexio_camera_a18,
        flexio_lcd_header,
    )

    print(f"FlexIO pins: {len(flexio_pins)}")
    print(f"EZH camera header rows: {len(ezh_camera)}")
    print(f"FLEXIO_LCD A18 camera rows: {len(flexio_camera_a18)}")
    print(f"FLEXIO_LCD header rows: {len(flexio_lcd_header)}")
    print(f"Output: {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
